"""
MODULE 2: Food Image Recognition + Nutrition Extraction
--------------------------------------------------------
Takes a photo of any food (packaged or unpackaged), uses Claude Vision API to:
  1. Identify the food item
  2. Estimate macro + micro nutrients per 100g
  3. Assign a confidence score (0-1)
  4. Output in the same standardized format as Module 1
Usage:
  python3 module2_image_nutrition.py --image /path/to/food.jpg
  python3 module2_image_nutrition.py --image /path/to/food.jpg --append  # add to module1_output.xlsx
Set your API key first:
  export ANTHROPIC_API_KEY="sk-ant-..."
Output columns match Module 1 exactly:
  Food_Name, Source_Kid, Serving_Size_g, Data_Type, Notes,
  Protein_g, Carbohydrate_g, Fat_g, SatFat_g, TransFat_g,
  TotalSugar_g, AddedSugar_g, Cholesterol_mg, Sodium_mg, Energy_kcal
"""
import os
import sys
import json
import base64
import argparse
import re
from pathlib import Path
from datetime import datetime
import anthropic
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
# ── Constants ─────────────────────────────────────────────────────────────────
NUTRIENT_KEYS = [
    # Macros
    "Protein_g", "Carbohydrate_g", "Fiber_g", "Fat_g",
    "SatFat_g", "TransFat_g", "TotalSugar_g", "AddedSugar_g",
    "Cholesterol_mg", "Sodium_mg", "Energy_kcal",
    # Vitamins
    "VitaminC_mg", "VitaminA_mcg", "VitaminK_mcg", "Folate_mcg",
    # Minerals
    "Calcium_mg", "Iron_mg", "Potassium_mg", "Magnesium_mg",
]
MODULE1_OUTPUT = "module1_output.xlsx"
MODULE2_OUTPUT = "module2_output.xlsx"
# The structured prompt that tells Claude exactly what to return
NUTRITION_PROMPT = """You are a nutrition expert assistant for a school science project.
Analyze this food image and return ONLY a JSON object with the following structure.
Do not include any explanation, markdown, or text outside the JSON.
{
  "food_name": "exact name of the food item",
  "food_type": "packaged" or "unpackaged/fresh",
  "serving_size_g": typical serving size in grams as a number,
  "confidence": confidence score from 0.0 to 1.0 (how confident you are in your nutrient estimates),
  "confidence_reason": "brief reason for confidence level",
  "nutrients_per_100g": {
    "Protein_g": <number>,
    "Carbohydrate_g": <number>,
    "Fiber_g": <number>,
    "Fat_g": <number>,
    "SatFat_g": <number>,
    "TransFat_g": <number>,
    "TotalSugar_g": <number>,
    "AddedSugar_g": <number>,
    "Cholesterol_mg": <number>,
    "Sodium_mg": <number>,
    "Energy_kcal": <number>,
    "VitaminC_mg": <number>,
    "VitaminA_mcg": <number>,
    "VitaminK_mcg": <number>,
    "Folate_mcg": <number>,
    "Calcium_mg": <number>,
    "Iron_mg": <number>,
    "Potassium_mg": <number>,
    "Magnesium_mg": <number>
  }
}
Rules:
- All nutrient values must be per 100g of the food
- If you cannot identify the food clearly, set confidence below 0.4
- Use standard nutritional database values (USDA / IFCT) — especially IFCT for Indian foods
- For packaged food where you can read the label, use label values (confidence will be higher)
- For homemade or mixed dishes, estimate based on typical recipe composition
- Never return null for nutrient values — use 0 if truly absent (e.g. TransFat in fresh fruit)
- Units: _g = grams, _mg = milligrams, _mcg = micrograms
- Fiber is especially important for fresh fruits and vegetables — always provide a value
- VitaminA_mcg is in RAE (Retinol Activity Equivalents)
"""
# ── Core Functions ─────────────────────────────────────────────────────────────
def load_image_as_base64(image_path: str) -> tuple[str, str]:
    """Load image and return (base64_data, media_type)."""
    path = Path(image_path)
    if not path.exists():
        raise FileNotFoundError(f"Image not found: {image_path}")
    ext = path.suffix.lower()
    media_type_map = {
        ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
        ".png": "image/png", ".gif": "image/gif",
        ".webp": "image/webp"
    }
    media_type = media_type_map.get(ext)
    if not media_type:
        raise ValueError(f"Unsupported image format: {ext}. Use JPG, PNG, GIF, or WEBP.")
    with open(image_path, "rb") as f:
        b64 = base64.standard_b64encode(f.read()).decode("utf-8")
    return b64, media_type
def call_claude_vision(image_path: str, api_key: str) -> dict:
    """Send image to Claude Vision API and parse the nutrition JSON response."""
    client = anthropic.Anthropic(api_key=api_key)
    print(f"  → Loading image: {image_path}")
    b64_data, media_type = load_image_as_base64(image_path)
    print(f"  → Calling Claude Vision API...")
    response = client.messages.create(
        model="claude-opus-4-6",
        max_tokens=1024,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": media_type,
                            "data": b64_data,
                        },
                    },
                    {
                        "type": "text",
                        "text": NUTRITION_PROMPT
                    }
                ],
            }
        ],
    )
    raw_text = response.content[0].text.strip()
    # Strip markdown code fences if present
    raw_text = re.sub(r"^```(?:json)?\s*", "", raw_text)
    raw_text = re.sub(r"\s*```$", "", raw_text)
    try:
        data = json.loads(raw_text)
    except json.JSONDecodeError as e:
        raise ValueError(f"Claude returned invalid JSON:\n{raw_text}\n\nError: {e}")
    return data
def build_record(api_response: dict, image_path: str, kid_name: str = "Image_Scan") -> dict:
    """Convert Claude API response into a standardized Module 1-compatible record."""
    nutrients = api_response.get("nutrients_per_100g", {})
    confidence = api_response.get("confidence", 0)
    conf_reason = api_response.get("confidence_reason", "")
    food_type = api_response.get("food_type", "unpackaged")
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    notes = (
        f"Source: Module 2 (Image scan) | "
        f"Image: {Path(image_path).name} | "
        f"Confidence: {confidence:.0%} ({conf_reason}) | "
        f"Scanned: {timestamp}"
    )
    record = {
        "Food_Name":      api_response.get("food_name", "Unknown Food"),
        "Source_Kid":     kid_name,
        "Serving_Size_g": api_response.get("serving_size_g"),
        "Data_Type":      "Packaged (label read)" if food_type == "packaged" else "Unpackaged (AI estimate)",
        "Notes":          notes,
        "Confidence":     confidence,  # extra column for Module 2
    }
    for key in NUTRIENT_KEYS:
        record[key] = nutrients.get(key)
    return record
def print_result(record: dict):
    """Pretty-print the result to console."""
    print(f"\n{'─'*60}")
    print(f"  Food identified: {record['Food_Name']}")
    print(f"  Data type:       {record['Data_Type']}")
    print(f"  Confidence:      {record.get('Confidence', 0):.0%}")
    print(f"  Serving size:    {record.get('Serving_Size_g', 'N/A')}g")
    print(f"\n  Nutrients per 100g:")
    nutrient_labels = {
        "Protein_g": "Protein", "Carbohydrate_g": "Carbohydrate", "Fat_g": "Total Fat",
        "SatFat_g": "Saturated Fat", "TransFat_g": "Trans Fat", "TotalSugar_g": "Total Sugar",
        "AddedSugar_g": "Added Sugar", "Cholesterol_mg": "Cholesterol",
        "Sodium_mg": "Sodium", "Energy_kcal": "Energy"
    }
    units = {
        "Protein_g": "g", "Carbohydrate_g": "g", "Fat_g": "g",
        "SatFat_g": "g", "TransFat_g": "g", "TotalSugar_g": "g",
        "AddedSugar_g": "g", "Cholesterol_mg": "mg", "Sodium_mg": "mg", "Energy_kcal": "kcal"
    }
    for key in NUTRIENT_KEYS:
        val = record.get(key)
        label = nutrient_labels.get(key, key)
        unit = units.get(key, "")
        val_str = f"{val:.1f} {unit}" if val is not None else "N/A"
        print(f"    {label:<20} {val_str}")
    print(f"{'─'*60}\n")
    if record.get("Confidence", 1) < 0.5:
        print("  WARNING: LOW CONFIDENCE — AI estimate may be inaccurate.")
        print("      Consider verifying with a nutrition label or database.\n")
def save_to_excel(records: list[dict], output_path: str):
    """Save Module 2 results to a formatted Excel file."""
    df = pd.DataFrame(records)
    # Reorder columns — put Confidence after Notes
    cols = ["Food_Name", "Source_Kid", "Serving_Size_g", "Data_Type",
            "Confidence", "Notes"] + NUTRIENT_KEYS
    df = df[[c for c in cols if c in df.columns]]
    wb = Workbook()
    ws = wb.active
    ws.title = "Module2_Image_Scans"
    hdr_fill = PatternFill("solid", fgColor="375623")  # green theme for Module 2
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    fill_a   = PatternFill("solid", fgColor="E2EFDA")
    fill_b   = PatternFill("solid", fgColor="FFFFFF")
    data_font = Font(name="Arial", size=10)
    headers = list(df.columns)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for ri, row in enumerate(df.itertuples(index=False), 2):
        fill = fill_a if ri % 2 == 0 else fill_b
        for ci, val in enumerate(row, 1):
            safe = None if (isinstance(val, float) and pd.isna(val)) else val
            c = ws.cell(ri, ci, safe)
            c.font = data_font
            c.fill = fill
            col_name = headers[ci - 1]
            if col_name in NUTRIENT_KEYS:
                c.number_format = "0.00"
                c.alignment = Alignment(horizontal="right")
            elif col_name == "Confidence":
                c.number_format = "0%"
                c.alignment = Alignment(horizontal="center")
            else:
                c.alignment = Alignment(horizontal="left")
    col_widths = {
        "Food_Name": 30, "Source_Kid": 14, "Serving_Size_g": 14,
        "Data_Type": 22, "Confidence": 13, "Notes": 60
    }
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(1, ci).column_letter].width = col_widths.get(h, 15)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)
    print(f"Saved to {output_path}")
def append_to_module1(record: dict, module1_path: str):
    """Append the new record to the existing module1_output.xlsx as a new row."""
    if not Path(module1_path).exists():
        print(f"  WARNING: {module1_path} not found — skipping append.")
        return
    wb = load_workbook(module1_path)
    ws = wb["Module1_Nutrition_Raw"]
    # Get headers from row 1
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    new_row = []
    for h in headers:
        val = record.get(h)
        new_row.append(None if (isinstance(val, float) and pd.isna(val)) else val)
    ws.append(new_row)
    # Apply light green fill to mark it as Module 2 data
    new_row_idx = ws.max_row
    fill = PatternFill("solid", fgColor="E2EFDA")
    for ci in range(1, len(headers) + 1):
        ws.cell(new_row_idx, ci).fill = fill
    wb.save(module1_path)
    print(f"Appended to {module1_path} (row {new_row_idx})")
# ── CLI Entry Point ────────────────────────────────────────────────────────────
def run(image_path: str, kid_name: str = "Image_Scan",
        append: bool = False, api_key: str = None):
    if not api_key:
        api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("\nERROR: No API key found.")
        print("   Set it with:  export ANTHROPIC_API_KEY='sk-ant-...'")
        print("   Or pass it:   run(..., api_key='sk-ant-...')\n")
        sys.exit(1)
    print(f"\nModule 2: Analyzing food image...")
    api_response = call_claude_vision(image_path, api_key)
    record = build_record(api_response, image_path, kid_name)
    print_result(record)
    # Save standalone Module 2 output
    save_to_excel([record], MODULE2_OUTPUT)
    # Optionally append to Module 1 master sheet
    if append:
        print(f"\nAppending to {MODULE1_OUTPUT}...")
        append_to_module1(record, MODULE1_OUTPUT)
    return record
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Module 2: Food Image -> Nutrition")
    parser.add_argument("--image",  required=True, help="Path to food image (JPG/PNG/WEBP)")
    parser.add_argument("--kid",    default="Image_Scan", help="Name of student scanning the food")
    parser.add_argument("--append", action="store_true",
                        help="Append result to module1_output.xlsx")
    parser.add_argument("--apikey", default=None,
                        help="Anthropic API key (or set ANTHROPIC_API_KEY env var)")
    args = parser.parse_args()
    run(
        image_path=args.image,
        kid_name=args.kid,
        append=args.append,
        api_key=args.apikey
    )
